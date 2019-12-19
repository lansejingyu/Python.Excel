#使用openpyxl模块对xlsx文件进行读操作

'''
xlrd和xlwt都是针对Excel97-2003操作的，也就是以xls结尾的文件。
Excel2007以上的版本，以xlsx为后缀。要对这种类型的Excel文件进行操作要使用openpyxl，该模块既可以进行“读”操作，也可以进行“写”操作，还可以对已经存在的文件做修改。
'''

import openpyxl

#获取工作簿对象
workbook = openpyxl.load_workbook("C:\\Users\\EDZ\\Desktop\\1.xlsx")
#与xlrd 模块的区别
#wokrbook=xlrd.open_workbook("C:\\Users\\EDZ\\Desktop\\1.xlsx")

#%%
#获取所有工作表名
#获取工作簿 workbook的所有工作表
shenames = workbook.get_sheet_names()
#在xlrd模块中为 sheetnames=workbook.sheet_names()
print(shenames)
#使用上述语(shenames = workbook.get_sheet_names())句会发出警告：DeprecationWarning: Call to deprecated function get_sheet_names (Use wb.sheetnames).
#说明 get_sheet_names已经被弃用 可以改用 wb.sheetnames 方法(25行)

#%%
shenames = workbook.sheetnames
print(shenames)

#%%
#获取工作表对象
worksheet = workbook.get_sheet_by_name('Sheet1')
print(worksheet)
#使用上述语句同样弹出警告：DeprecationWarning: Call to deprecated function get_sheet_by_name (Use wb[sheetname]).
#说明改写成如下格式get_sheet_by_name已经被弃用 可以改用wb[sheetname]（34行）

#%%
worksheet=workbook["Sheet1"]
print(worksheet)

#还可以通过如下写法获得表对象
#%%
#还可以通过如下写法获得表对象
worksheet = workbook[shenames[0]]
print(worksheet)

#%%
#根据索引方式获取工作表对象
#还可以通过索引方式获取表对象
worksheet = workbook.worksheets[0]
print(worksheet)

#%%
#获取工作表的属性
name = worksheet.title   # #获取表名   在xlrd中为worksheet.name
print(name)

#%%
#获取该表相应的行数和列数
rows = worksheet.max_row
columns = worksheet.max_column
print(rows,columns)
#在xlrd中为 worksheet.nrows  worksheet.ncols

#%%
#按行或列方式获取表中的数据
'''
要想以行方式或者列方式，获取整个工作表的内容，我们需要使用到以下两个生成器：
sheet.rows，这是一个生成器，里面是每一行数据，每一行数据由一个元组类型包裹。
sheet.columns，同上，里面是每一列数据。
'''
